import imaplib
import email
from email.header import decode_header
from transformers import pipeline
import openpyxl
import os
from tqdm import tqdm
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import smtplib
from email.message import EmailMessage
import re

class EmailToneAnalyzer:
    def __init__(self, imap_server, email_account, email_password, mailbox="INBOX"):
        # Email configuration
        self.imap_server = imap_server
        self.email_account = email_account
        self.email_password = email_password
        self.mailbox = mailbox
        
        # Set up transformer model for sentiment analysis
        print("Loading sentiment analysis model...")
        self.sentiment_pipeline = pipeline(
            "text-classification", 
            model="SamLowe/roberta-base-go_emotions",  # Better for detecting emotions in text
            top_k=None  # Return all emotions with scores
        )
        
        # Create Excel file
        self.today_str = datetime.now().strftime("%Y-%m-%d")
        self.excel_file = f"email_tone_{self.today_str}.xlsx"
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.append(["Sender", "Subject", "Date", "Primary Emotion", "Score", "Secondary Emotions"])
        
        # Apply formatting 
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            self.ws.column_dimensions[col].width = 25
    
    def clean_text(self, text):
        """Clean text for better sentiment analysis"""
        if not text:
            return ""
        
        # Remove HTML tags
        text = re.sub(r'<[^>]+>', ' ', text)
        
        # Remove email signatures
        text = re.sub(r'--+\s*\n.*', '', text, flags=re.DOTALL)
        
        # Remove URLs
        text = re.sub(r'http\S+', '', text)
        
        # Remove extra whitespace
        text = re.sub(r'\s+', ' ', text).strip()
        
        return text
    
    def analyze_tone(self, text):
        """Analyze the emotional tone of text using transformer model"""
        if not text or len(text.strip()) < 5:
            return "NEUTRAL", 0.5, {}
        
        cleaned_text = self.clean_text(text)
        
        # For very long texts, analyze only the first 1024 tokens
        # Most important emotional content tends to be at the beginning
        max_length = 1024
        
        try:
            results = self.sentiment_pipeline(cleaned_text[:max_length])
            if not results or not results[0]:
                return "NEUTRAL", 0.5, {}
            
            # Get all emotions with scores
            emotions = {item['label']: item['score'] for item in results[0]}
            
            # Find primary emotion (highest score)
            primary_emotion = max(emotions.items(), key=lambda x: x[1])
            
            # Get secondary emotions (emotions with scores > 0.1)
            secondary_emotions = {k: round(v, 3) for k, v in emotions.items() 
                                 if v > 0.1 and k != primary_emotion[0]}
            
            return primary_emotion[0].upper(), round(primary_emotion[1], 3), secondary_emotions
            
        except Exception as e:
            print(f"Error analyzing tone: {e}")
            return "ERROR", 0.0, {}
    
    def extract_email_body(self, msg):
        """Extract email body text, handling multipart messages"""
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                try:
                    if content_type == "text/plain":
                        body = part.get_payload(decode=True).decode(errors="ignore")
                        break
                    elif content_type == "text/html" and not body:
                        html = part.get_payload(decode=True).decode(errors="ignore")
                        soup = BeautifulSoup(html, "html.parser")
                        body = soup.get_text()
                except Exception as e:
                    print(f"Error parsing email part: {e}")
        else:
            content_type = msg.get_content_type()
            try:
                if content_type == "text/plain":
                    body = msg.get_payload(decode=True).decode(errors="ignore")
                elif content_type == "text/html":
                    html = msg.get_payload(decode=True).decode(errors="ignore")
                    soup = BeautifulSoup(html, "html.parser")
                    body = soup.get_text()
            except Exception as e:
                print(f"Error decoding email body: {e}")
        
        return body
    
    def fetch_and_analyze_emails(self, days_back=1):
        """Fetch emails from the last X days and analyze their tone"""
        try:
            # Connect to IMAP
            mail = imaplib.IMAP4_SSL(self.imap_server)
            mail.login(self.email_account, self.email_password)
            mail.select(self.mailbox)
            
            # Search emails from the specified time period
            date_since = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")
            status, messages = mail.search(None, f'(SINCE {date_since})')
            email_ids = messages[0].split()
            
            if not email_ids:
                print(f"No emails found since {date_since}")
                return
                
            print(f"Found {len(email_ids)} emails to analyze")
            
            # Process emails
            for eid in tqdm(email_ids, desc="Processing Emails", unit="email"):
                _, msg_data = mail.fetch(eid, "(RFC822)")
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])
                        
                        # Extract email metadata
                        subject, encoding = decode_header(msg["Subject"])[0]
                        if isinstance(subject, bytes):
                            subject = subject.decode(encoding if encoding else "utf-8")
                        
                        from_ = msg.get("From")
                        date_ = msg.get("Date")
                        
                        # Extract body
                        body = self.extract_email_body(msg)
                        
                        # Analyze tone
                        primary_emotion, score, secondary_emotions = self.analyze_tone(body)
                        
                        # Format secondary emotions for Excel
                        secondary_emotions_text = ", ".join([f"{k}: {v}" for k, v in secondary_emotions.items()])
                        
                        # Truncate body for Excel preview
                        preview = body[:300] + "..." if len(body) > 300 else body
                        preview = preview.replace("\n", " ").strip()
                        
                        # Sanitize text for Excel to prevent formula and special character issues
                        def sanitize_for_excel(text):
                            if not text:
                                return ""
                            # Remove null bytes and replace problematic characters
                            text = text.replace('\x00', '').replace('\r', '').replace('\t', ' ')
                            # Remove non-ASCII characters
                            text = text.encode("ascii", "ignore").decode()
                            # Truncate excessively long strings
                            max_length = 32767
                            if len(text) > max_length:
                                text = text[:max_length - 3] + "..."
                            # Add single quotes to prevent formula interpretation
                            if isinstance(text, str) and text and text[0] in ['=', '+', '-', '@']:
                                text = "'" + text
                            return text
                        
                        # Sanitize all text fields
                        safe_from = sanitize_for_excel(from_)
                        safe_subject = sanitize_for_excel(subject)
                        safe_date = sanitize_for_excel(date_)
                        safe_emotion = sanitize_for_excel(primary_emotion)
                        safe_secondary = sanitize_for_excel(secondary_emotions_text)
                        
                        # Append to Excel with sanitized values
                        self.ws.append([
                            safe_from, 
                            safe_subject, 
                            safe_date, 
                            safe_emotion, 
                            score,  # Numeric value doesn't need sanitizing
                            safe_secondary,
                        ])
            
            # Close the connection
            mail.logout()
            
        except Exception as e:
            print(f"Error fetching emails: {e}")
    
    def send_report_via_email(self, to_email):
        """Send the Excel report via email"""
        # Save Excel file first
        self.wb.save(self.excel_file)
        
        from_email = self.email_account
        from_password = self.email_password
        
        msg = EmailMessage()
        msg["Subject"] = f"Daily Email Sentiment Analysis Report - {self.today_str}"
        msg["From"] = from_email
        msg["To"] = to_email
        
        # Add summary content
        total_emails = self.ws.max_row - 1  # Subtract header row
        
        # Count each primary emotion type
        emotion_counts = {}
        for row in range(2, self.ws.max_row + 1):
            emotion = self.ws.cell(row=row, column=4).value
            emotion_counts[emotion] = emotion_counts.get(emotion, 0) + 1
        
        # Format emotion summary
        emotion_summary = "\n".join([
            f"- {emotion}: {count} emails ({round(count/total_emails*100, 1)}%)" 
            for emotion, count in sorted(emotion_counts.items(), key=lambda x: x[1], reverse=True)
        ])
        
        content = f"""Dear Sir Bilal,

Attached is the daily email sentiment analysis report for {self.today_str}.

Summary:

- Total emails analyzed: {total_emails}

{emotion_summary}

The analysis identifies emotions present in each email, which can help prioritize and categorize incoming communications.
Please review the attached Excel file for detailed information.

This is an automated message.

Best Regards,

🕶️ FBI (Feelings & Behavior Investigator) 
We read between the lines — and also judge them.
"""
        msg.set_content(content)
        
        # Attach the Excel file
        with open(self.excel_file, "rb") as f:
            file_data = f.read()
            file_name = os.path.basename(self.excel_file)
            msg.add_attachment(
                file_data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=file_name
            )
        
        try:
            server = smtplib.SMTP_SSL(self.imap_server, 465)
            server.login(from_email, from_password)
            server.send_message(msg)
            server.quit()
            print(f"📧 Report emailed successfully to {to_email}")
        except Exception as e:
            print(f"❌ Failed to send email: {e}")
    
    def run_analysis(self, days_back=1, recipient_email=None):
        """Run the complete analysis workflow"""
        print(f"🔍 Starting email tone analysis for the past {days_back} days...")
        self.fetch_and_analyze_emails(days_back)
        self.wb.save(self.excel_file)
        print(f"💾 Analysis saved to {self.excel_file}")
        
        if recipient_email:
            self.send_report_via_email(recipient_email)
        
        print("✅ Email tone analysis complete!")


# Usage example
if __name__ == "__main__":
    # Configuration
    IMAP_SERVER = "mailserver"
    EMAIL_ACCOUNT = "sender@example.com"
    EMAIL_PASSWORD = "Password"
    MAILBOX = "mailbox"
    
    # Create analyzer and run
    analyzer = EmailToneAnalyzer(
        imap_server=IMAP_SERVER,
        email_account=EMAIL_ACCOUNT,
        email_password=EMAIL_PASSWORD,
        mailbox=MAILBOX
    )
    
    analyzer.run_analysis(days_back=1, recipient_email="recipient@example.com")
    
